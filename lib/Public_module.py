import os
import logging, time
import pymysql
import pymssql
import copy
from openpyxl import load_workbook

# 当前文件路径
cur_path = os.path.dirname(os.path.realpath(__file__))
print(cur_path)

# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 类名称：InsertLog_P
# 类的目的：写日志
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：谢殿俊
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
class InsertLog_P():
    def __init__(self):
        # log_path是存放日志的路径
        log_path = os.path.join(os.path.dirname(cur_path), 'log')
        print(log_path)
        # 如果不存在这个logs文件夹，就自动创建一个
        if not os.path.exists(log_path):
            os.mkdir(log_path)
        # 文件的命名
        self.logname = os.path.join(log_path, '%s.log' %time.strftime('%Y_%m_%d'))
        self.logger = logging.getLogger()
        self.logger.setLevel(logging.DEBUG)
        # 日志输出格式
        self.formatter = logging.Formatter('[%(asctime)s - %(funcName)s line: %(lineno)3d] - %(levelname)s: %(message)s')

    def __console(self, level, message):
        # 创建一个FileHandler，用于写到本地
        # fh = logging.FileHandler(self.logname, 'a')  # 追加模式  这个是python2的
        fh = logging.FileHandler(self.logname, 'a', encoding='utf-8')  # 这个是python3的
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(self.formatter)
        self.logger.addHandler(fh)

        # 创建一个StreamHandler,用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        ch.setFormatter(self.formatter)
        self.logger.addHandler(ch)

        if level == 'info':
            self.logger.info(message)
        elif level == 'debug':
            self.logger.debug(message)
        elif level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        # 这两行代码是为了避免日志输出重复问题
        self.logger.removeHandler(ch)
        self.logger.removeHandler(fh)
        # 关闭打开的文件
        fh.close()

    def debug(self, message):
        self.__console('debug', message)

    def info(self, message):
        self.__console('info', message)

    def warning(self, message):
        self.__console('warning', message)

    def error(self, message):
        self.__console('error', message)


# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 函数/过程名称：GetSkipScripts_P
# 函数/过程的目的：获取不需要执行的模块名字
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：谢殿俊
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------

def GetSkipScripts_P(FilePath):
	try:
		skipscript = []
		wb = load_workbook(FilePath)
		ws = wb['ScriptPath']
		# print(ws)
		rowcount = ws.max_row
		# print(rowcount)
		for i in range(1, rowcount + 1):
			cellvalue = ws.cell(row=i, column=3).value
			# print(cellvalue)
			if cellvalue == 'False':
				modulename = ws.cell(row=i, column=2).value
				# print(modulename)
				skipscript.append(modulename)
		wb.close()
		return skipscript
	except BaseException as msg:
		log = InsertLog_P()
		log.error(msg)


#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# 函数/过程名称：GetSkipTestCases_P
# 函数/过程的目的：获取不需要执行用例名字
# 假设：无
# 影响：无
# 输入：无
# 返回值：无
# 创建者：谢殿俊
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
#-------------------------------------------------------------------------------

def GetSkipTestCases_P(FilePath):
	try:
		skiptestcase=[]
		wb=load_workbook(FilePath)
		sheets=wb.sheetnames
		# print(sheets)
		for i in sheets:
			ws=wb[i]
			rowcount=ws.max_row
			for j in range(2,rowcount+1):
				cellvalue=ws.cell(row=j,column=7).value
				# print(cellvalue)
				if cellvalue =='False':
					casename=ws.cell(row=j,column=1).value
					skiptestcase.append(casename)
		wb.close()
		return skiptestcase
	except BaseException as msg:
		log=InsertLog_P()
		log.error(msg)


# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 函数/过程名称：Connection_MySQLData
# 函数/过程的目的：操作MYSQL数据库
# 假设：无
# 影响：无
# 输入：无
# 返回值：删除数据库数据（元组类型）
# 创建者：谢殿俊
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
class Connection_MySQLData():
	#打开数据库连接
	def Open(self):
		host = 'localhost'      # 数据库地址'localhost',3306,'edu','edu','edu'
		port = 3306             # 端口
		user = 'edu'            # 用户名
		password = 'edu'        # 密码
		db = 'edu'              # 数据库名称
		try:
			self.conn = pymysql.connect(host=host,port=port,user=user,passwd=password,db=db)
			self.cursor = self.conn.cursor()
		except BaseException as msg:
			log = InsertLog_P()
			log.error(msg)
		
	# 关闭数据库连接
	def Close(self):
		self.cursor.close()
		self.conn.close()
		
	# 新建表
	def CreateTable(self, sql):
		self.Open()
		self.cursor.execute(sql)
		self.conn.commit()
		self.Close()
	
	# 添加数据
	def InsertData(self, sql):
		self.Open()
		result=self.cursor.executemany(sql)
		self.conn.commit()
		self.Close()
		return result
	# 删除数据
	def DeleteData(self, sql):
		self.Open()
		result=self.cursor.execute(sql)
		self.conn.commit()
		self.Close()
		return result
	# 查询语句
	def SelectTable(self, sql):
		self.Open()
		self.cursor.execute(sql)
		result = self.cursor.fetchall()
		self.Close()
		return result
	
	# 修改语句
	def UpdateData(self, sql):
		self.Open()
		result=self.cursor.execute(sql)
		self.conn.commit()
		self.Close()
		return result
# -------------------------------------------------------------------------------
# -------------------------------------------------------------------------------
# 函数/过程名称：Connection_Sqlsever_Data
# 函数/过程的目的：操作sqlserver数据库
# 假设：无
# 影响：无
# 输入：无
# 返回值：（元组类型）
# 创建者：谢殿俊
# 创建时间：2019/05/18
# 修改者：
# 修改原因：
# 修改时间:
# -------------------------------------------------------------------------------
class Connection_Sqlsever_Data():
	def __init__(self):
		self.server = "143.35.208.112"      # server     数据库服务器名称或IP
		self.user = "sa"                    # user       用户名
		self.password = "sinoworld"         # password   密码
		self.database = "E_pmp"             #database   数据库名称
	# 数据库远程连接
	def Open(self):
		try:
			self.conn = pymssql.connect(self.server, self.user, self.password, self.database)
			# 使用cursor()方法获取操作游标
			self.cursor =self.conn.cursor()
		except BaseException as msg:
			log = InsertLog_P()
			log.error(msg)
	#关闭数据库连接
	def Close(self):
		self.cursor.close()
		self.conn.close()
	#新建表
	def CreateTable(self,sql):
		self.Open()
		self.cursor.execute(sql)
		self.conn.commit()
		self.Close()
	#添加数据
	def InsertData(self,sql):
		self.Open()
		self.cursor.executemany(sql)
		self.conn.commit()
		self.Close()
		
	#删除数据
	def DeleteData(self,sql):
		self.Open()
		self.cursor.execute(sql)
		self.conn.commit()
		self.Close()
	# 查询语句
	def SelectTable(self,sql):
		self.Open()
		self.cursor.execute(sql)
		result = self.cursor.fetchall()
		self.Close()
		return result
	#修改语句
	def UpdateData(self,sql):
		self.Open()
		self.cursor.execute(sql)
		self.conn.commit()
		self.Close()
		


def get_test_suite(discover,m,t):
    #筛选出并去除不需要执行的脚本
    suite_m = copy.deepcopy(discover)
    # print(suite_m)
    for i in range(len(m)):
        for j in range(discover._tests.__len__()):
            d = discover._tests[j]
            if m[i] in str(d):
                suite_m._tests.remove(d)
    #筛选出并去除不需要执行的用例
    suite_c = copy.deepcopy(suite_m)
    for i in range(len(t)):
        for j in range(suite_m._tests.__len__()):
            s_m =  suite_m._tests[j]
            for z in range(s_m._tests.__len__()):
                s_c = s_m._tests[z]
                for k in range(s_c._tests.__len__()):
                    s_t = s_c._tests[k]
                    if t[i] == s_t._testMethodName:
                        suite_c._tests[j]._tests[z]._tests.remove(s_t)
    return suite_c
if __name__ == '__main__':
	# ConfigFilePath = "../po/config.xlsx"
	# GetSkipScripts_P(ConfigFilePath)
	# r='../testcase/测试用例.xlsx'
	# GetSkipTestCases_P(r)
	# r = Connection_MySQLData()
	# r.Open()
	# r.DeleteMySQLData_P('localhost',3306,'edu','edu','edu')
	# sql="delete from xsmart_users where phone='13811112222'"
	# result=r.DeleteData(sql)
	# print(result)
	InsertLog_P()